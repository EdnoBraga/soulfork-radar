"""Exportação: CSV e XLSX."""
from __future__ import annotations

import csv
from pathlib import Path

_PERIGO = ("=", "+", "-", "@", "\t", "\r")


def _seguro(v):
    """Neutraliza injeção de fórmula: célula que começa com =,+,-,@ vira texto."""
    if isinstance(v, str) and v and v.startswith(_PERIGO):
        return "'" + v
    return v

from .models import Lead

COLUNAS = [
    ("score", "Score"),
    ("faixa", "Faixa"),
    ("nome", "Empresa"),
    ("nicho", "Nicho"),
    ("municipio", "Município"),
    ("uf", "UF"),
    ("telefone", "Telefone"),
    ("whatsapp_link", "WhatsApp"),
    ("email", "E-mail principal"),
    ("emails_extra", "Outros e-mails"),
    ("site", "Site"),
    ("instagram", "Instagram"),
    ("instagram_url", "URL Instagram"),
    ("tiktok", "TikTok"),
    ("facebook", "Facebook"),
    ("linkedin", "LinkedIn"),
    ("nota", "Nota Google"),
    ("avaliacoes", "Avaliações"),
    ("cnpj", "CNPJ"),
    ("razao_social", "Razão social"),
    ("situacao_cadastral", "Situação cadastral"),
    ("cnae", "CNAE"),
    ("porte", "Porte"),
    ("abertura", "Abertura"),
    ("site_no_ar", "Site no ar"),
    ("https", "HTTPS"),
    ("responsivo", "Responsivo"),
    ("tempo_ms", "Tempo (ms)"),
    ("gtm", "GTM"),
    ("pixel", "Meta Pixel"),
    ("ga4", "GA4"),
    ("politica", "Política LGPD"),
    ("formulario", "Formulário"),
    ("cms", "CMS"),
    ("achados", "Achados"),
    ("endereco", "Endereço"),
    ("google_maps_url", "Google Maps"),
    ("coletado_em", "Coletado em"),
]


def _sim_nao(v) -> str:
    if v is True:
        return "sim"
    if v is False:
        return "não"
    return "não verificável"


def achatar(lead: Lead) -> dict:
    c, r, d = lead.contatos, lead.redes, lead.diagnostico
    achados = " | ".join(
        f"{s.titulo}" for s in lead.sinais if s.pontos > 0
    )
    return {
        "score": lead.score,
        "faixa": lead.faixa,
        "nome": lead.nome,
        "nicho": lead.nicho,
        "municipio": lead.municipio,
        "uf": lead.uf,
        "telefone": c.telefone or "",
        "whatsapp_link": c.whatsapp or "",
        "email": c.emails[0] if c.emails else "",
        "emails_extra": ", ".join(c.emails[1:]) if len(c.emails) > 1 else "",
        "site": lead.site or "",
        "instagram": f"@{r.instagram}" if r.instagram else "",
        "instagram_url": f"https://instagram.com/{r.instagram}" if r.instagram else "",
        "tiktok": f"@{r.tiktok}" if r.tiktok else "",
        "facebook": f"https://facebook.com/{r.facebook}" if r.facebook else "",
        "linkedin": f"https://linkedin.com/company/{r.linkedin}" if r.linkedin else "",
        "nota": lead.nota if lead.nota is not None else "",
        "avaliacoes": lead.avaliacoes if lead.avaliacoes is not None else "",
        "cnpj": lead.cnpj or "",
        "razao_social": lead.razao_social or "",
        "situacao_cadastral": lead.situacao_cadastral or "",
        "cnae": lead.cnae or "",
        "porte": lead.porte or "",
        "abertura": lead.abertura or "",
        "site_no_ar": _sim_nao(d.site_no_ar),
        "https": _sim_nao(d.https),
        "responsivo": _sim_nao(d.responsivo),
        "tempo_ms": d.tempo_resposta_ms or "",
        "gtm": _sim_nao(d.tem_gtm),
        "pixel": _sim_nao(d.tem_meta_pixel),
        "ga4": "encontrado" if d.tem_ga4 else "não verificável",
        "politica": ("quebrada" if d.politica_quebrada else _sim_nao(d.tem_politica_privacidade)),
        "formulario": _sim_nao(d.tem_formulario),
        "cms": f"{d.cms or ''} {d.cms_versao or ''}".strip(),
        "achados": achados,
        "endereco": lead.endereco or "",
        "google_maps_url": lead.google_maps_url or "",
        "coletado_em": lead.coletado_em or "",
    }


def para_csv(leads: list[Lead], destino: str | Path) -> Path:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([rotulo for _, rotulo in COLUNAS])
        for lead in leads:
            linha = achatar(lead)
            w.writerow([_seguro(linha.get(k, "")) for k, _ in COLUNAS])
    return destino


def para_xlsx(leads: list[Lead], destino: str | Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    cab_fill = PatternFill("solid", fgColor="1F2937")
    cab_font = Font(color="FFFFFF", bold=True, size=10)
    cores = {"quente": "FEE2E2", "morno": "FEF3C7", "frio": "E0F2FE", "fraco": "F3F4F6",
             "descartar": "E5E7EB"}

    ws.append([rotulo for _, rotulo in COLUNAS])
    for c in ws[1]:
        c.fill = cab_fill
        c.font = cab_font
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "C2"

    for lead in leads:
        linha = achatar(lead)
        ws.append([_seguro(linha.get(k, "")) for k, _ in COLUNAS])
        fill = PatternFill("solid", fgColor=cores.get(lead.faixa, "FFFFFF"))
        for c in ws[ws.max_row]:
            c.fill = fill

    larguras = {"Empresa": 34, "Achados": 70, "Site": 32, "Endereço": 42,
                "E-mail principal": 30, "Outros e-mails": 30, "Razão social": 34,
                "CNAE": 40, "URL Instagram": 32, "Google Maps": 24, "WhatsApp": 16}
    for i, (_, rotulo) in enumerate(COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(rotulo, 14)

    ws.auto_filter.ref = ws.dimensions
    wb.save(destino)
    return destino
